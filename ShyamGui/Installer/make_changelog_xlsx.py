"""Generate a styled Excel changelog (CHANGELOG.xlsx) for the
Atomik Acoustic Simulation Engine, mirroring CHANGELOG.md."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Atomik palette --------------------------------------------------------
BG      = "0D0D12"   # near-black background
PANEL   = "17171D"   # panel
ACCENT  = "44D9E6"   # cyan
HEAD    = "F5F7FA"   # near-white
LABEL   = "B8C0CC"   # soft gray
SUCCESS = "53D769"
WARNING = "FF6B6B"
ZEBRA   = "1E1E26"   # alternate row

thin = Side(style="thin", color="2C2F38")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def title_font(size=18, color=HEAD):
    return Font(name="Calibri", size=size, bold=True, color=color)


def head_font():
    return Font(name="Calibri", size=11, bold=True, color=BG)


def body_font(color=HEAD, bold=False):
    return Font(name="Calibri", size=10, color=color, bold=bold)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.font = head_font()
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 22


def write_table(ws, start_row, headers, rows, widths, type_colors=None,
                type_col=None):
    # header
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)
    style_header_row(ws, start_row, len(headers))

    # body
    r = start_row + 1
    for ridx, row in enumerate(rows):
        for i, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = BORDER
            base = ZEBRA if ridx % 2 else PANEL
            cell.fill = PatternFill("solid", fgColor=base)
            cell.font = body_font(LABEL if i != 1 else HEAD)
        # colour the Type column if requested
        if type_colors and type_col:
            t = row[type_col - 1]
            col = type_colors.get(t)
            if col:
                tc = ws.cell(row=r, column=type_col)
                tc.font = body_font(col, bold=True)
        r += 1

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return r


def banner(ws, text, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = title_font(18, HEAD)
    c.fill = PatternFill("solid", fgColor=BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=sub)
    s.font = Font(name="Calibri", size=10, italic=True, color=ACCENT)
    s.fill = PatternFill("solid", fgColor=BG)
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18


def paint_bg(ws, nrows, ncols):
    for r in range(1, nrows + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = PatternFill("solid", fgColor=BG)


wb = Workbook()

# ===========================================================================
# Sheet 1: Overview
# ===========================================================================
ws = wb.active
ws.title = "Overview"
ws.sheet_view.showGridLines = False
banner(ws, "Atomik Acoustic Simulation Engine - Changelog",
       "Generated from CHANGELOG.md  |  follows Keep a Changelog + SemVer", 4)

rows = [
    ["1.1.0", "2026-06-29", "Major feature + UX release",
     "Theming, project dashboard/.atmk, heatmap + PDF export, grid/CAD-DXF, "
     "array presets, dual datasets, responsive window, branded installer"],
    ["1.0.0", "2026-06-22", "Initial baseline",
     "Dual-subwoofer (XN18) prediction & visualisation in a 30x30 m world; "
     "measured directivity added 2026-06-24"],
]
write_table(ws, 4, ["Version", "Date", "Summary", "Highlights"],
            rows, [12, 14, 26, 80])
ws.freeze_panes = "A5"

# ===========================================================================
# Sheet 2: v1.1.0 detailed changes
# ===========================================================================
ws2 = wb.create_sheet("v1.1.0 Changes")
ws2.sheet_view.showGridLines = False
banner(ws2, "v1.1.0  -  Detailed Changes",
       "Development window 2026-06-25 -> 2026-06-29", 5)

A, C, F = "Added", "Changed", "Fixed"
changes = [
    # Section, Type, Item, Details, Date
    ["Theming & Preferences", A, "Dark / Light themes",
     "Dynamic app-wide switching via theme-aware palette + AtomikLookAndFeel; every screen/graph/dialog/heatmap chrome adapts live", "2026-06-25"],
    ["Theming & Preferences", A, "Preferences / Settings tab",
     "Choose theme and unit system (SI / Imperial); displayed values convert, internal calcs stay SI", "2026-06-26"],
    ["Theming & Preferences", A, "Persistent preferences",
     "Theme, units, grid visibility, measurement set, recent projects stored via PropertiesFile; broadcast via ChangeBroadcaster", "2026-06-26"],

    ["Project Workflow", A, "Project Dashboard at launch",
     "Separate window: new / open / recent projects; then opens main simulation window", "2026-06-26"],
    ["Project Workflow", A, "Project metadata capture",
     "Name, Engineer, Owner, Address, City, Country, Email, Mobile, Date - used in exports/reports", "2026-06-26"],
    ["Project Workflow", A, "Single-file .atmk projects",
     "Self-contained JSON: metadata + scene (speakers) + simulation settings; save/open", "2026-06-26"],

    ["Heatmap Export", A, "Client-ready heatmap sheet",
     "Header (logo, software name, title), project details, heatmap data (frequency, scale, units, legend) - generation logic untouched", "2026-06-26"],

    ["PDF Report", A, "Multi-page PDF report",
     "Cover, Project Info, Input Parameters, Results, Graphs, Heatmaps (30/50/80/100 Hz), Summary - generated programmatically", "2026-06-26"],
    ["PDF Report", A, "Custom PDF 1.4 writer (PdfDocument)",
     "Vector text + embedded images, no external dependency", "2026-06-26"],
    ["PDF Report", A, "AcousticAnalysis module",
     "RT60 (Sabine), absorption, transmission loss (mass-law), frequency response sampled from engine", "2026-06-26"],
    ["PDF Report", A, "GraphRender",
     "Line/bar charts rendered to images for the report", "2026-06-26"],

    ["Workspace / Grid / Layout", A, "Improved grid",
     "Major/minor lines, zoom-adaptive spacing (niceStep), higher contrast; show/hide toggle (persisted)", "2026-06-26"],
    ["Workspace / Grid / Layout", A, "Layout reference import",
     "Floor plans as background references - images (PNG/JPG) and DXF drawings", "2026-06-26"],
    ["Workspace / Grid / Layout", A, "DXF parser (DxfImport)",
     "Dependency-free 2D parser: LINE, LWPOLYLINE, POLYLINE/VERTEX, CIRCLE, ARC", "2026-06-26"],
    ["Workspace / Grid / Layout", A, "Layout alignment & visibility tools",
     "Move, rotate, scale (width in m), opacity, show/hide, lock, snap-to-grid; live drag", "2026-06-26"],

    ["Array Presets", A, "Cardioid presets (2/3/4 sub)",
     "Ready-to-use cardioid subwoofer arrays", "2026-06-26"],
    ["Array Presets", A, "End-fired presets (2/3/4 sub)",
     "Auto-calculated spacing, delay, polarity, orientation; manually editable after placement", "2026-06-26"],

    ["Measurement Datasets", A, "Two selectable sets (kept separate)",
     "Ground Plane = factory open-field (30/60/100/150/200 Hz, 1 m); Room = GYLT (30/80/200/500 Hz, 0.5/1 m)", "2026-06-29"],
    ["Measurement Datasets", A, "Per-set heatmaps & polar plots",
     "Switching reloads only that dataset + rebuilds directivity; each polar normalised to its own on-axis level; selection persisted", "2026-06-29"],
    ["Measurement Datasets", A, "Dynamic availability/legend",
     "Measured-polar 'available frequencies' and legend generated from the loaded dataset", "2026-06-29"],

    ["Responsive UI", A, "Minimum window size",
     "Main 1120x686, dashboard 680x537 - layout can never collapse/hide controls", "2026-06-27"],
    ["Responsive UI", A, "Fixed widescreen aspect ratio",
     "1340:820 preserved during resize - no stretching distortion", "2026-06-27"],
    ["Responsive UI", A, "Defensive layout clamps",
     "Workspace visible (P1), side panels usable (P2), controls keep usable size (P3)", "2026-06-27"],
    ["Responsive UI", A, "Scrollable control panel",
     "Dense controls scroll (viewport) instead of shrinking", "2026-06-27"],
    ["Responsive UI", A, "Responsive header + 'more options' menu",
     "Title auto-shrinks/never clips; Export PDF / Save / Preferences fold into kebab menu at all sizes", "2026-06-26"],

    ["Branding & Installer", A, "Atomik app icon embedded",
     "Windows .rc icon for exe, installer wizard, Add/Remove Programs, shortcuts; generated by make_icon.py", "2026-06-26"],
    ["Branding & Installer", A, "Windows installer (Inno Setup)",
     "Single Setup.exe: app + both datasets + fonts + icon; static CRT (no VC++ redist); per-user; uninstaller", "2026-06-26"],

    ["Versioning", C, "1.0.0 -> 1.1.0",
     "App title shows v1.1; installer 1.1", "2026-06-26"],
    ["Color system", C, "Removed all red; cyan accent system",
     "Header red->near-white; accents red->grey->cyan; hierarchy headings #F5F7FA / labels #B8C0CC / muted #7E8794; bg #0D0D12 / panel #17171D; accent #44D9E6; success #53D769; warning #FF6B6B; PDF FR line red->steel-blue", "2026-06-26"],
    ["Measured directivity", C, "Data-driven normalisation",
     "Each curve normalised to its own on-axis level instead of hard-coded reference SPLs", "2026-06-29"],
    ["Measurement Datasets", C, "Option renames",
     "Open Field -> Ground Plane; GYLT -> Room (datasets unchanged)", "2026-06-29"],

    ["Header", F, "Buttons no longer overlap title",
     "Fixed overlap in narrow/short windows", "2026-06-26"],
    ["Header", F, "Title no longer clips",
     "Was showing 'v1.' instead of 'v1.1'", "2026-06-26"],
    ["Recent Projects", F, "Recent Projects hygiene",
     "Entries whose .atmk no longer exists are hidden and pruned automatically", "2026-06-27"],
    ["Installer", F, "Fresh install guarantee",
     "Per-user settings folder cleared on install + uninstall; no stale recent projects/settings on reinstall", "2026-06-27"],
]

type_colors = {A: SUCCESS, C: ACCENT, F: WARNING}
write_table(ws2, 4, ["Section", "Type", "Item", "Details", "Date"],
            changes, [24, 10, 34, 70, 13],
            type_colors=type_colors, type_col=2)
ws2.freeze_panes = "A5"
ws2.auto_filter.ref = f"A4:E{4 + len(changes)}"

# ===========================================================================
# Sheet 3: v1.0.0 baseline
# ===========================================================================
ws3 = wb.create_sheet("v1.0.0 Baseline")
ws3.sheet_view.showGridLines = False
banner(ws3, "v1.0.0  -  Initial Baseline",
       "2026-06-22  (measured directivity added 2026-06-24)", 3)

base = [
    ["View modes", "SPL Heatmap, Pressure Map, Interference, Directivity, Measured Polar, Fit View", ""],
    ["Acoustic metrics", "SPL, Frequency Response, RT60, Absorption, Transmission Loss, heatmaps", ""],
    ["Frequency selector", "20, 30, 40, 50, 60, 70, 80, 90, 100, 110, 120, 150, 200, 500 Hz", ""],
    ["Measured directivity", "Horizontal directivity from .xlsx; applied when sim frequency matches; 1/3-octave smoothing; optional 3 dB contour bands", "2026-06-24"],
    ["Speaker / scene controls", "Per-unit X/Y, gain (dB), delay (ms), invert polarity, reverse orientation, enable/disable; add/delete; quick layouts (1/2/3 devices)", ""],
    ["Simulation settings", "Grid resolution, dB floor", ""],
    ["Exports", "PNG image, CSV data, SVG figure", ""],
    ["Engine", "Monopole summation + frequency-dependent piston directivity fallback; fractional-octave band power averaging", ""],
    ["Branding & UI", "JUCE desktop app; Montserrat + Space Mono fonts; Atomik identity", ""],
]
write_table(ws3, 4, ["Area", "Description", "Date"], base, [24, 90, 13])
ws3.freeze_panes = "A5"

# ===========================================================================
# Sheet 4: Timeline
# ===========================================================================
ws4 = wb.create_sheet("Timeline")
ws4.sheet_view.showGridLines = False
banner(ws4, "Development Timeline",
       "What landed when (derived from source history)", 2)

timeline = [
    ["2026-06-22", "v1.0 baseline: engine + UI (Main, MainComponent, ControlPanel, InfoPanel, RadiationPatternComponent, AcousticEngine, ColourMaps)"],
    ["2026-06-24", "Measured-directivity loading added (MeasurementData)"],
    ["2026-06-25", "v1.1 theming foundation: theme-aware palette + AtomikLookAndFeel (BrandTheme)"],
    ["2026-06-26", "Preferences/settings, .atmk projects + dashboard, enhanced heatmap export, PDF pipeline, grid + CAD/DXF, array presets, app icon + installer, header responsiveness, version bump, red->cyan overhaul"],
    ["2026-06-27", "Window scaling (min size, fixed aspect ratio, clamps); fresh-install guarantee; Recent Projects hygiene"],
    ["2026-06-29", "Dual measurement datasets (Ground Plane / Room), selectable + separate; final naming"],
]
write_table(ws4, 4, ["Date", "Work landed"], timeline, [14, 110])
ws4.freeze_panes = "A5"

# paint remaining backgrounds for a clean dark look
for sheet in (ws, ws2, ws3, ws4):
    paint_bg(sheet, sheet.max_row + 2, sheet.max_column)

out = os.path.join(os.path.dirname(__file__), "..", "CHANGELOG.xlsx")
out = os.path.abspath(out)
wb.save(out)
print("Wrote", out)
