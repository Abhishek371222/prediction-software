"""Deep analysis of all polar measurement Excel files across 3 folders."""
import json
import re
from pathlib import Path

import numpy as np
import openpyxl

ROOT = Path(r"d:\prediction software")
CACHE = Path(r"C:\Users\User\.cursor\projects\d-prediction-software\canvases\_analysis_cache.json")


def load_xlsx(path):
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0] if rows else None
    ang, spl = [], []
    for i, row in enumerate(rows):
        if i == 0:
            continue
        if row[0] is None or row[1] is None:
            continue
        ang.append(float(row[0]))
        spl.append(float(row[1]))
    ang = np.array(ang)
    spl = np.array(spl)
    if len(ang):
        o = np.argsort(ang)
        ang, spl = ang[o], spl[o]
    return {
        "headers": headers,
        "ang": ang,
        "spl": spl,
        "sheet_title": ws.title,
        "max_col": ws.max_column,
        "max_row": ws.max_row,
    }


def parse_name(name):
    m = re.match(
        r"Freq(?:uency|unecy)_(\d+(?:\.\d+)?)_(\d+(?:\.\d+)?)Horizantal\.xlsx",
        name,
        re.I,
    )
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


def wrap_angle(a):
    return ((a + 180) % 360) - 180


def stats(ang, spl):
    if len(ang) == 0:
        return None
    a = wrap_angle(ang)
    i0 = int(np.argmin(np.abs(a)))
    on_axis = float(spl[i0])
    peak = float(spl.max())
    peak_ang = float(ang[int(np.argmax(spl))])
    trough = float(spl.min())
    trough_ang = float(ang[int(np.argmin(spl))])
    spl_n = spl - peak
    spl_n0 = spl - on_axis
    front = spl_n[np.abs(a) <= 90]
    rear = spl_n[np.abs(a) > 90]
    fb = float(front.mean() - rear.mean()) if len(front) and len(rear) else None

    idx = np.argsort(a)
    a_s, s_s = a[idx], spl_n[idx]
    zero = int(np.argmin(np.abs(a_s)))

    def beamwidth(thresh):
        lo = a_s[zero]
        for k in range(zero, -1, -1):
            if s_s[k] >= thresh:
                lo = a_s[k]
            else:
                break
        hi = a_s[zero]
        for k in range(zero, len(a_s)):
            if s_s[k] >= thresh:
                hi = a_s[k]
            else:
                break
        return float(hi - lo)

    bw6 = beamwidth(-6)
    bw3 = beamwidth(-3)
    p2 = 10 ** (spl_n / 10.0)
    di = float(-10 * np.log10(p2.mean()))
    left = spl_n[(a < 0) & (a > -180)]
    right = spl_n[(a > 0) & (a < 180)]
    asym = float(left.mean() - right.mean()) if len(left) and len(right) else None
    i180 = int(np.argmin(np.minimum(np.abs(a - 180), np.abs(a + 180))))
    rear_level = float(spl_n[i180])
    roughness = float(np.mean(np.abs(np.diff(spl, n=2)))) if len(spl) >= 3 else None
    steps = np.diff(ang)
    step_unique = sorted(set(np.round(steps, 6).tolist())) if len(steps) else []
    eavg = float(10 * np.log10(np.mean(10 ** (spl / 10.0))))
    cardinals = {}
    for target in [0, 30, 60, 90, 120, 150, 180, 210, 240, 270, 300, 330]:
        ia = int(np.argmin(np.abs(ang - target)))
        cardinals[target] = float(spl[ia])
    points = [
        {
            "deg": float(ang[i]),
            "spl": float(spl[i]),
            "rel_peak": float(spl_n[i]),
            "rel_onaxis": float(spl_n0[i]),
        }
        for i in range(len(ang))
    ]
    return {
        "n_points": int(len(ang)),
        "angle_min": float(ang.min()),
        "angle_max": float(ang.max()),
        "angle_steps_unique": step_unique,
        "on_axis_spl": on_axis,
        "peak_spl": peak,
        "peak_angle": peak_ang,
        "trough_spl": trough,
        "trough_angle": trough_ang,
        "dynamic_range": float(peak - trough),
        "energy_avg_spl": eavg,
        "fb_ratio_db": fb,
        "beamwidth_6db": bw6,
        "beamwidth_3db": bw3,
        "di_2d_db": di,
        "asymmetry_L_minus_R_db": asym,
        "rear_level_rel_peak_db": rear_level,
        "roughness_mean_abs_d2": roughness,
        "spl_mean": float(spl.mean()),
        "spl_std": float(spl.std()),
        "spl_median": float(np.median(spl)),
        "cardinals": cardinals,
        "points": points,
        "peak_minus_onaxis": float(peak - on_axis),
    }


datasets = {
    "3_inch_frequency_sweep": list((ROOT / "3 inch frequency sweep").glob("*.xlsx")),
    "Factory_Readings": list((ROOT / "Factory Readings").rglob("*.xlsx")),
    "shyamGuildMeasurements": list((ROOT / "shyamGuildMeasurements").glob("*.xlsx")),
}

all_results = {}
for ds, files in datasets.items():
    all_results[ds] = {}
    for f in sorted(files):
        freq, dist = parse_name(f.name)
        data = load_xlsx(f)
        st = stats(data["ang"], data["spl"])
        all_results[ds][f.name] = {
            "path": str(f),
            "freq_hz": freq,
            "distance_m": dist,
            "headers": [str(h) if h is not None else None for h in (data["headers"] or [])],
            "sheet": data["sheet_title"],
            "max_col": data["max_col"],
            "max_row": data["max_row"],
            "stats": st,
        }

for ds, items in all_results.items():
    print("\n" + "=" * 90)
    print(f"DATASET: {ds}  ({len(items)} files)")
    print("=" * 90)
    for key, rec in sorted(
        items.items(), key=lambda kv: (kv[1]["freq_hz"] or 0, kv[1]["distance_m"] or 0)
    ):
        st = rec["stats"]
        print(f"\n--- {key} ---")
        print(
            f"  freq={rec['freq_hz']} Hz  dist={rec['distance_m']} m  "
            f"sheet={rec['sheet']}  cols={rec['max_col']} rows={rec['max_row']}"
        )
        print(f"  headers={rec['headers']}")
        if not st:
            print("  EMPTY")
            continue
        print(
            f"  points={st['n_points']}  angles={st['angle_min']}..{st['angle_max']}  "
            f"steps={st['angle_steps_unique']}"
        )
        print(
            f"  on-axis={st['on_axis_spl']:.3f} dB  peak={st['peak_spl']:.3f} dB "
            f"@{st['peak_angle']:.0f}°  trough={st['trough_spl']:.3f} dB "
            f"@{st['trough_angle']:.0f}°"
        )
        print(
            f"  peak-onaxis={st['peak_minus_onaxis']:.3f}  dyn_range={st['dynamic_range']:.3f}  "
            f"eavg={st['energy_avg_spl']:.3f}  mean={st['spl_mean']:.3f}  "
            f"std={st['spl_std']:.3f}  med={st['spl_median']:.3f}"
        )
        print(
            f"  -6dB BW={st['beamwidth_6db']:.0f}°  -3dB BW={st['beamwidth_3db']:.0f}°  "
            f"F/B={st['fb_ratio_db']:.2f} dB  DI~={st['di_2d_db']:.2f} dB"
        )
        print(
            f"  asym(L-R)={st['asymmetry_L_minus_R_db']:.3f} dB  "
            f"rear@180={st['rear_level_rel_peak_db']:.2f} dB  "
            f"roughness={st['roughness_mean_abs_d2']:.4f}"
        )
        print("  FULL SWEEP (deg, SPL, rel_peak, rel_onaxis):")
        for p in st["points"]:
            print(
                f"    {p['deg']:6.1f}  {p['spl']:8.3f}  "
                f"{p['rel_peak']:7.2f}  {p['rel_onaxis']:7.2f}"
            )

print("\n" + "=" * 90)
print("CROSS-DATASET COMPARISONS")
print("=" * 90)

lookup = {}
for ds, items in all_results.items():
    for key, rec in items.items():
        if rec["freq_hz"] is None:
            continue
        lookup[(ds, rec["freq_hz"], rec["distance_m"])] = rec

print("\n--- ShyamGuild vs Factory at 1 m (overlapping freqs) ---")
for freq in [30, 200]:
    s = lookup.get(("shyamGuildMeasurements", float(freq), 1.0))
    f = lookup.get(("Factory_Readings", float(freq), 1.0))
    if s and f:
        ss, fs = s["stats"], f["stats"]
        print(f"\n{freq} Hz @ 1m:")
        print(
            f"  Shyam: on-axis={ss['on_axis_spl']:.3f} peak={ss['peak_spl']:.3f} "
            f"dyn={ss['dynamic_range']:.3f} BW6={ss['beamwidth_6db']:.0f} "
            f"FB={ss['fb_ratio_db']:.2f} DI={ss['di_2d_db']:.2f}"
        )
        print(
            f"  Fact : on-axis={fs['on_axis_spl']:.3f} peak={fs['peak_spl']:.3f} "
            f"dyn={fs['dynamic_range']:.3f} BW6={fs['beamwidth_6db']:.0f} "
            f"FB={fs['fb_ratio_db']:.2f} DI={fs['di_2d_db']:.2f}"
        )
        print(f"  Delta on-axis (Shyam-Fact) = {ss['on_axis_spl'] - fs['on_axis_spl']:.3f} dB")
        sa = {p["deg"]: p["spl"] for p in ss["points"]}
        fa = {p["deg"]: p["spl"] for p in fs["points"]}
        common = sorted(set(sa) & set(fa))
        diffs = [sa[d] - fa[d] for d in common]
        print(
            f"  Pointwise SPL delta (Shyam-Fact): mean={np.mean(diffs):.3f} "
            f"std={np.std(diffs):.3f} min={np.min(diffs):.3f} max={np.max(diffs):.3f}"
        )
        s_n = np.array([sa[d] for d in common])
        s_n = s_n - s_n.max()
        f_n = np.array([fa[d] for d in common])
        f_n = f_n - f_n.max()
        corr = np.corrcoef(s_n, f_n)[0, 1]
        rmse = np.sqrt(np.mean((s_n - f_n) ** 2))
        print(f"  Shape corr (norm to peak)={corr:.4f}  RMSE={rmse:.3f} dB")

print("\n--- ShyamGuild distance law (0.5 / 1 / 2 m) ---")
for freq in [30, 80, 200, 500]:
    rows = []
    for dist in [0.5, 1.0, 2.0]:
        rec = lookup.get(("shyamGuildMeasurements", float(freq), dist))
        if rec:
            st = rec["stats"]
            rows.append(
                (
                    dist,
                    st["on_axis_spl"],
                    st["peak_spl"],
                    st["energy_avg_spl"],
                    st["beamwidth_6db"],
                    st["di_2d_db"],
                    st["fb_ratio_db"],
                )
            )
    print(f"\n{freq} Hz:")
    print(f"  {'dist':>5} {'onaxis':>8} {'peak':>8} {'eavg':>8} {'BW6':>6} {'DI':>6} {'FB':>6}")
    for r in rows:
        print(
            f"  {r[0]:>5.1f} {r[1]:>8.3f} {r[2]:>8.3f} {r[3]:>8.3f} "
            f"{r[4]:>6.0f} {r[5]:>6.2f} {r[6]:>6.2f}"
        )
    for i in range(len(rows) - 1):
        d1, oa1 = rows[i][0], rows[i][1]
        d2, oa2 = rows[i + 1][0], rows[i + 1][1]
        expected = -20 * np.log10(d2 / d1)
        actual = oa2 - oa1
        print(
            f"  {d1}->{d2}m: actual dOn-axis={actual:.2f} dB, "
            f"free-field expected={expected:.2f} dB, excess={actual - expected:.2f} dB"
        )

print("\n--- 3 inch frequency sweep trends @ 1 m ---")
print(
    f"  {'freq':>7} {'onaxis':>8} {'peak':>8} {'dyn':>7} {'BW6':>6} "
    f"{'BW3':>6} {'FB':>7} {'DI':>6} {'rear':>7} {'asym':>7}"
)
for freq in [300, 500, 1000, 2000, 4000, 8000, 16000]:
    rec = lookup.get(("3_inch_frequency_sweep", float(freq), 1.0))
    if not rec:
        continue
    st = rec["stats"]
    print(
        f"  {freq:>7} {st['on_axis_spl']:>8.3f} {st['peak_spl']:>8.3f} "
        f"{st['dynamic_range']:>7.2f} {st['beamwidth_6db']:>6.0f} "
        f"{st['beamwidth_3db']:>6.0f} {st['fb_ratio_db']:>7.2f} "
        f"{st['di_2d_db']:>6.2f} {st['rear_level_rel_peak_db']:>7.2f} "
        f"{st['asymmetry_L_minus_R_db']:>7.3f}"
    )

print("\n--- Factory Readings trends @ 1 m ---")
print(
    f"  {'freq':>7} {'onaxis':>8} {'peak':>8} {'dyn':>7} {'BW6':>6} "
    f"{'BW3':>6} {'FB':>7} {'DI':>6} {'rear':>7} {'asym':>7}"
)
for freq in [30, 60, 100, 150, 200]:
    rec = lookup.get(("Factory_Readings", float(freq), 1.0))
    if not rec:
        continue
    st = rec["stats"]
    print(
        f"  {freq:>7} {st['on_axis_spl']:>8.3f} {st['peak_spl']:>8.3f} "
        f"{st['dynamic_range']:>7.2f} {st['beamwidth_6db']:>6.0f} "
        f"{st['beamwidth_3db']:>6.0f} {st['fb_ratio_db']:>7.2f} "
        f"{st['di_2d_db']:>6.2f} {st['rear_level_rel_peak_db']:>7.2f} "
        f"{st['asymmetry_L_minus_R_db']:>7.3f}"
    )

print("\n--- 500 Hz: ShyamGuild vs 3-inch sweep @ 1 m ---")
s = lookup.get(("shyamGuildMeasurements", 500.0, 1.0))
t = lookup.get(("3_inch_frequency_sweep", 500.0, 1.0))
if s and t:
    ss, ts = s["stats"], t["stats"]
    print(
        f"  Shyam: on-axis={ss['on_axis_spl']:.3f} peak={ss['peak_spl']:.3f} "
        f"dyn={ss['dynamic_range']:.3f} BW6={ss['beamwidth_6db']:.0f} "
        f"FB={ss['fb_ratio_db']:.2f} DI={ss['di_2d_db']:.2f}"
    )
    print(
        f"  3inch: on-axis={ts['on_axis_spl']:.3f} peak={ts['peak_spl']:.3f} "
        f"dyn={ts['dynamic_range']:.3f} BW6={ts['beamwidth_6db']:.0f} "
        f"FB={ts['fb_ratio_db']:.2f} DI={ts['di_2d_db']:.2f}"
    )
    sa = {p["deg"]: p["spl"] for p in ss["points"]}
    ta = {p["deg"]: p["spl"] for p in ts["points"]}
    common = sorted(set(sa) & set(ta))
    s_n = np.array([sa[d] for d in common])
    s_n = s_n - s_n.max()
    t_n = np.array([ta[d] for d in common])
    t_n = t_n - t_n.max()
    corr = np.corrcoef(s_n, t_n)[0, 1]
    rmse = np.sqrt(np.mean((s_n - t_n) ** 2))
    print(
        f"  Shape corr={corr:.4f} RMSE={rmse:.3f} dB  "
        f"dOn-axis={ss['on_axis_spl'] - ts['on_axis_spl']:.3f}"
    )

print("\n" + "=" * 90)
print("DATA QUALITY AUDIT")
print("=" * 90)
for ds, items in all_results.items():
    for key, rec in items.items():
        st = rec["stats"]
        issues = []
        if not st:
            issues.append("EMPTY")
        else:
            if st["n_points"] != 37:
                issues.append(f"unexpected point count {st['n_points']} (expect 37 for 0..360/10)")
            if st["angle_min"] != 0 or st["angle_max"] != 360:
                issues.append(f"angle range {st['angle_min']}-{st['angle_max']}")
            steps = [s for s in st["angle_steps_unique"] if abs(s - 10) > 1e-6]
            if steps:
                issues.append(f"non-10deg steps: {st['angle_steps_unique']}")
            pts = {p["deg"]: p["spl"] for p in st["points"]}
            if 0 in pts and 360 in pts and abs(pts[0] - pts[360]) > 0.01:
                issues.append(f"0° vs 360° mismatch: {pts[0]:.3f} vs {pts[360]:.3f}")
            if st["roughness_mean_abs_d2"] and st["roughness_mean_abs_d2"] > 2.0:
                issues.append(f"high roughness {st['roughness_mean_abs_d2']:.2f}")
            if abs(st["asymmetry_L_minus_R_db"]) > 1.5:
                issues.append(f"strong L/R asymmetry {st['asymmetry_L_minus_R_db']:.2f} dB")
            if st["peak_minus_onaxis"] > 0.5:
                issues.append(
                    f"peak not on-axis (offset {st['peak_minus_onaxis']:.2f} dB "
                    f"@ {st['peak_angle']:.0f}°)"
                )
        status = "OK" if not issues else "; ".join(issues)
        print(f"  [{ds}] {key}: {status}")

out = {}
for ds, items in all_results.items():
    out[ds] = []
    for key, rec in items.items():
        st = rec["stats"]
        if not st:
            continue
        out[ds].append(
            {
                "file": key,
                "freq": rec["freq_hz"],
                "dist": rec["distance_m"],
                "on_axis": st["on_axis_spl"],
                "peak": st["peak_spl"],
                "peak_ang": st["peak_angle"],
                "trough": st["trough_spl"],
                "trough_ang": st["trough_angle"],
                "dyn": st["dynamic_range"],
                "eavg": st["energy_avg_spl"],
                "bw6": st["beamwidth_6db"],
                "bw3": st["beamwidth_3db"],
                "fb": st["fb_ratio_db"],
                "di": st["di_2d_db"],
                "asym": st["asymmetry_L_minus_R_db"],
                "rear": st["rear_level_rel_peak_db"],
                "rough": st["roughness_mean_abs_d2"],
                "std": st["spl_std"],
                "cardinals": {str(k): v for k, v in st["cardinals"].items()},
                "points": st["points"],
            }
        )

CACHE.parent.mkdir(parents=True, exist_ok=True)
with open(CACHE, "w", encoding="utf-8") as f:
    json.dump(out, f)
print(f"\nWrote cache: {CACHE}")
print(f"Total files analyzed: {sum(len(v) for v in all_results.values())}")
