#!/usr/bin/env python3
"""Export Q21S BEM → pack polar CSVs. L↔R mirror only; native rear from 10 m field."""
from __future__ import annotations

import math
from collections import defaultdict
from pathlib import Path

import numpy as np
import openpyxl
from scipy.interpolate import NearestNDInterpolator, griddata

ROOT = Path(__file__).resolve().parents[2]
# Prefer new 10 m BEM folder
CANDIDATES = [
    ROOT / "BEM_Data_10m" / "Q21S_10m_PolarPlotData.xlsx",
    ROOT / "Q21S_PolarPlot_Data_10M.xlsx",
]
OUT = ROOT / "ShyamGui/prediction software/MeasurementIntegrationPack/Data"
SOURCE_OUT = OUT / "source" / "Q21S"

# Native BEM_Data_10m/<Hz>Hz.xlsx only — no interpolated extras.
CATALOGUE = [20, 29, 52, 81, 98, 153, 198, 256, 309, 352, 400, 401]
DISTANCES = [0.5, 1.0, 2.0]
P_REF = 20e-6
C0 = 343.0
PISTON_R = 0.13


def main():
    xlsx = next((p for p in CANDIDATES if p.exists()), None)
    if xlsx is None:
        raise SystemExit("No BEM xlsx found")
    OUT.mkdir(parents=True, exist_ok=True)
    SOURCE_OUT.mkdir(parents=True, exist_ok=True)

    print("Loading", xlsx)
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    by_f = defaultdict(list)
    sheets = [s for s in wb.sheetnames if s.lower() != "explanation"]
    for sname in sheets:
        for row in wb[sname].iter_rows(values_only=True):
            if not row or row[0] is None or isinstance(row[0], str):
                continue
            try:
                x, z = float(row[1]), float(row[3])
                f = float(row[4]); pr, pi = float(row[5]), float(row[6])
            except Exception:
                continue
            by_f[f].append((x, z, pr, pi))
    wb.close()
    bem_freqs = np.array(sorted(by_f.keys()), dtype=float)
    print(f"freqs={len(bem_freqs)} pts@first={len(by_f[float(bem_freqs[0])])}")

    def spl_db(pr, pi):
        m = math.hypot(pr, pi)
        return -200.0 if m <= 0 else 20.0 * math.log10(m / P_REF)

    def mirror_lr(samples):
        pts = {}
        for x, z, pr, pi in samples:
            s = spl_db(pr, pi)
            for sx in (x, -x):
                key = (round(sx, 8), round(z, 8))
                pts[key] = s if key not in pts else 0.5 * (pts[key] + s)
        return (np.array([k[0] for k in pts]),
                np.array([k[1] for k in pts]),
                np.array([pts[k] for k in pts]))

    def polar_at_radius(xs, zs, ss, radius, hz, n=360):
        angles = np.arange(0, n, dtype=float)
        qx = radius * np.sin(np.deg2rad(angles))
        qz = radius * np.cos(np.deg2rad(angles))
        vals = griddata((xs, zs), ss, (qx, qz), method="linear")
        miss = np.isnan(vals)
        if miss.any():
            nn = NearestNDInterpolator(np.column_stack([xs, zs]), ss)
            vals[miss] = nn(qx[miss], qz[miss])

        # If native rear exists (zmin < 0), keep it. Else taper rear only.
        has_rear = float(np.min(zs)) < -0.1
        if has_rear:
            return angles, vals

        ka = (2 * math.pi * max(hz, 1.0) / C0) * PISTON_R
        sigma = max(1.0, ka)
        def taper(w_deg):
            w = math.radians(abs(w_deg))
            return 1.0 / (1.0 + math.exp(sigma * (w - math.pi / 2.0)))
        t_side = max(taper(90.0), 1e-6)
        for i, a in enumerate(angles):
            w = a if a <= 180.0 else a - 360.0
            if abs(w) <= 90.0:
                continue
            side = vals[90] if w > 0 else vals[270]
            lin = (10 ** (float(side) / 20.0)) * (taper(w) / t_side)
            vals[i] = 20.0 * math.log10(max(lin, 1e-30))
        return angles, vals

    fields = {float(f): mirror_lr(by_f[float(f)]) for f in bem_freqs}

    def anchors(hz):
        if hz <= bem_freqs[0]:
            return fields[float(bem_freqs[0])], None, 0.0
        if hz >= bem_freqs[-1]:
            return fields[float(bem_freqs[-1])], None, 0.0
        i = int(np.searchsorted(bem_freqs, hz))
        f0, f1 = float(bem_freqs[i - 1]), float(bem_freqs[i])
        if abs(f0 - hz) < 1e-9:
            return fields[f0], None, 0.0
        if abs(f1 - hz) < 1e-9:
            return fields[f1], None, 0.0
        a = math.log(max(f0, 1e-9)); b = math.log(max(f1, 1e-9)); x = math.log(max(hz, 1e-9))
        return (f0, f1), (fields[f0], fields[f1]), (x - a) / max(b - a, 1e-12)

    old = []
    man = OUT / "manifest.csv"
    if man.exists():
        for line in man.read_text().splitlines():
            if line.startswith("set,") or line.startswith("Q21S,") or not line.strip():
                continue
            old.append(line)
    body = ["set,freq_hz,distance_m,file,source_xlsx,n_points"] + old

    for hz in CATALOGUE:
        left, right, t = anchors(float(hz))
        for dist in DISTANCES:
            if right is None:
                xs, zs, ss = left
                ang, spl = polar_at_radius(xs, zs, ss, dist, float(hz))
            else:
                (xs0, zs0, ss0), (xs1, zs1, ss1) = right
                a0, s0 = polar_at_radius(xs0, zs0, ss0, dist, float(hz))
                _, s1 = polar_at_radius(xs1, zs1, ss1, dist, float(hz))
                p = (1 - t) * (10 ** (s0 / 20)) + t * (10 ** (s1 / 20))
                ang, spl = a0, 20 * np.log10(np.maximum(p, 1e-30))
            tag = {0.5: "0p5m", 1.0: "1p0m", 2.0: "2p0m"}[dist]
            fname = f"Q21S_{hz}Hz_{tag}.csv"
            with (OUT / fname).open("w") as f:
                f.write("degree,dBSPL\n")
                for a, s in zip(ang, spl):
                    f.write(f"{a:.1f},{s:.6f}\n")
            body.append(f"Q21S,{hz},{dist:.1f},{fname},{xlsx.name},360")
            print(f"  {fname} on={spl[0]:.1f} rear={spl[180]:.1f} dB")

    man.write_text("\n".join(body) + "\n")
    print("done", man)


if __name__ == "__main__":
    main()
